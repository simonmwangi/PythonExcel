# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/

import openpyxl  # to read the excel file

path = "files/gfg.xlsx"

# to open the workbook
wb_obj = openpyxl.load_workbook(path)

# to get the active sheet object
sheet_obj = wb_obj.active

####Note>>>The first row or column integer is 1 not 0

# To obtain the max column and max row of the active sheet
row = sheet_obj.max_row
column = sheet_obj.max_column

print("The total number of rows: ", row)
print("The total number of columns: ", column)

# Looping through the fist column
for i in range(1, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print(cell_obj.value)
# Printing the values of the third row
for i in range(1, column + 1):
    cell_obj = sheet_obj.cell(row=3, column=i)
    print(cell_obj.value, end=" ")

# Reading from multiple cells using the cell name
cell_obj = sheet_obj['A1': 'B6']

for cell1, cell2 in cell_obj:
    print(cell1.value, cell2.value)

# Reading from a single cell
cell_obj = sheet_obj.cell(row=1, column=1)
print(cell_obj.value)

####Writing To Cells
# Call the Workbook() function to create a new blank workbook project

from openpyxl import workbook  # to get the workbook function

workbook = openpyxl.Workbook()

# to get the active sheet from the workbook
sheet = workbook.active

# writing to specific row and column
c1 = sheet.cell(row=1, column=1)
c1.value = "HELLO"

c2 = sheet.cell(row=1, column=2)
c2.value = "WORLD-"

# Saving the workbook after write
workbook.save(filename="files/sample.xlsx")

new_path = "files/sample.xlsx"
workbook = openpyxl.load_workbook(new_path)
sheet = workbook.active
# m_cell = sheet["A1":"B2"]
# for a,b in m_cell:
#    print(a.value, b.value)
# print(m_cell)

####Appending to a workbook